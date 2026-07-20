#!/usr/bin/env python3
"""Converts xls/Scan.xlsx ('Scan App' sheet) into json/data/Scan.json.

Sheet layout (discovered by inspection, not assumed):
  Cols A-H  (1-8):   Version, source, Level 2..7 -- the feature outline, one row per UI element.
  Cols I-AZ (9-52):  44 printer-model columns. Header is 4 stacked rows:
                        row1 = product family (merged),  row3 = platform segment (merged),
                        row4 = model name,  row5 = engine class,  row6 = WIP/Ready status.
                      Family and segment colors vary per group (e.g. "Esnl Enhanced" is green,
                      "Workflow UI" is blue; "CISS" is yellow, "SOHO/LOW-END" is pink, etc) --
                      captured per-model via merge-aware fill resolution, not a single tab-wide
                      color like the simpler 2-Line IA sheet.
  Cols BA-BD (53-56): "Components: Setting row" -- per-row UI widget metadata (TextField/Button/...).
                      Band on row1 (coral), sub-headers "Level 2".."Level 5" on row4.
  Cols BF-BK (58-63): "Factory Quick Sets" -- per-row Default/- flags for Jupiter (& Beam MF) quicksets.
                      Band on row1 (purple), key label row2, product-line row3, target-model row4.
  Col BL (64):        EPICS/STORIES ticket references, per row. Band on row1 (bright purple).
  Col BM (65):        Design Notes, per row. Label on row4, no fill.
Real data ends at row 632; columns beyond BM are unused formatting artifacts.
"""
import json
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

SRC = "../xls/Scan.xlsx"
OUT = "json/data/Scan.json"
SCHEMA = "json/data/_schema.json"

MODEL_START, MODEL_END = column_index_from_string("I"), column_index_from_string("AZ")
COMPONENT_START, COMPONENT_END = column_index_from_string("BA"), column_index_from_string("BD")
QUICKSET_START, QUICKSET_END = column_index_from_string("BF"), column_index_from_string("BK")
EPIC_COL = column_index_from_string("BL")
NOTES_COL = column_index_from_string("BM")
LAST_ROW = 632


def merged_value(ws, row, col):
    """Returns the cell's value, resolving merged ranges to their anchor value."""
    cell = ws.cell(row=row, column=col)
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            return ws.cell(row=rng.min_row, column=rng.min_col).value
    return cell.value


def hex_of(cell):
    rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
    if rgb in (None, "00000000"):
        return None
    return f"#{rgb[2:]}" if len(rgb) == 8 else f"#{rgb}"


def merged_fill(ws, row, col):
    """Same anchor-resolution as merged_value, but for fill color -- merged
    cells only carry real style on their top-left anchor; every other cell
    in the range reports blank style even though it visually renders filled."""
    cell = ws.cell(row=row, column=col)
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            return hex_of(ws.cell(row=rng.min_row, column=rng.min_col))
    return hex_of(cell)


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def cell_style(cell):
    """Captures the original xlsx presentation for one cell -- fill color and
    bold -- so the frontend can reproduce the sheet's own visual tree
    language instead of re-deriving depth from scratch. Returns None for a
    plain/unstyled cell to keep the JSON small."""
    fill = hex_of(cell)
    bold = bool(cell.font.bold) if cell.font else False
    if fill is None and not bold:
        return None
    return {"fill": fill, "bold": bold}


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Scan App"]

    models = []
    for c in range(MODEL_START, MODEL_END + 1):
        name = clean(ws.cell(row=4, column=c).value)
        if not name:
            continue
        models.append({
            "key": name,
            "column": get_column_letter(c),
            "family": clean(merged_value(ws, 1, c)),
            "familyFill": merged_fill(ws, 1, c),
            "segment": clean(merged_value(ws, 3, c)),
            "segmentFill": merged_fill(ws, 3, c),
            "engineClass": clean(ws.cell(row=5, column=c).value),
            "status": clean(ws.cell(row=6, column=c).value),
        })

    component_headers = {c: clean(ws.cell(row=4, column=c).value) for c in range(COMPONENT_START, COMPONENT_END + 1)}
    quickset_headers = {}
    for c in range(QUICKSET_START, QUICKSET_END + 1):
        name = clean(ws.cell(row=2, column=c).value)
        if name:
            quickset_headers[c] = {
                "key": name,
                "line": clean(ws.cell(row=3, column=c).value),
                "lineFill": hex_of(ws.cell(row=3, column=c)),
                "model": clean(ws.cell(row=4, column=c).value),
            }

    level_field_names = ["level2", "level3", "level4", "level5", "level6", "level7"]
    style_field_names = ["version", "source"] + level_field_names

    rows = []
    for r in range(7, LAST_ROW + 1):
        level_vals = [clean(ws.cell(row=r, column=c).value) for c in range(1, 9)]
        model_vals = {m["key"]: clean(ws.cell(row=r, column=column_index_from_string(m["column"])).value) for m in models}
        component_vals = {v: clean(ws.cell(row=r, column=c).value) for c, v in component_headers.items() if v}
        quickset_vals = {v["key"]: clean(ws.cell(row=r, column=c).value) for c, v in quickset_headers.items()}
        epic = clean(ws.cell(row=r, column=EPIC_COL).value)
        notes = clean(ws.cell(row=r, column=NOTES_COL).value)

        row_has_content = any(level_vals) or any(model_vals.values()) or any(component_vals.values()) or any(quickset_vals.values()) or epic or notes
        if not row_has_content:
            continue

        entry = {
            "row": r,
            "version": level_vals[0],
            "source": level_vals[1],
            "level2": level_vals[2],
            "level3": level_vals[3],
            "level4": level_vals[4],
            "level5": level_vals[5],
            "level6": level_vals[6],
            "level7": level_vals[7],
        }
        if any(model_vals.values()):
            entry["models"] = {k: v for k, v in model_vals.items() if v is not None}
        if any(component_vals.values()):
            entry["componentSetting"] = component_vals
        if any(quickset_vals.values()):
            entry["quickSets"] = quickset_vals
        if epic:
            entry["epicStory"] = epic
        if notes:
            entry["designNotes"] = notes

        # Original xlsx presentation (fill color, bold) for the feature-tree
        # columns on this row -- lets the frontend render the sheet's own
        # tree/depth visual language (a bright-green highlight here, not the
        # peach used in 2-Line IA -- colors are captured per-sheet, never
        # assumed) instead of flattening it.
        cell_styles = {}
        for field_name, col in zip(style_field_names, range(1, 9)):
            style = cell_style(ws.cell(row=r, column=col))
            if style:
                cell_styles[field_name] = style
        if cell_styles:
            entry["cellStyle"] = cell_styles

        rows.append(entry)

    header_style = {
        "treeHeaderFill": hex_of(ws.cell(row=4, column=1)),
        "statusRowFill": hex_of(ws.cell(row=6, column=MODEL_START)),
        "componentsBandFill": hex_of(ws.cell(row=1, column=COMPONENT_START)),
        "componentsBandLabel": clean(ws.cell(row=1, column=COMPONENT_START).value),
        "quickSetsBandFill": hex_of(ws.cell(row=1, column=QUICKSET_START)),
        "quickSetsBandLabel": clean(ws.cell(row=1, column=QUICKSET_START).value),
        "epicBandFill": hex_of(ws.cell(row=1, column=EPIC_COL)),
        "epicLabel": clean(ws.cell(row=1, column=EPIC_COL).value),
        "notesLabel": clean(ws.cell(row=4, column=NOTES_COL).value),
    }

    feature_tree_labels = [clean(ws.cell(row=4, column=c).value) or n for n, c in zip(style_field_names, range(1, 9))]
    component_columns = [v for v in component_headers.values() if v]

    out = {
        "tab": "Scan",
        "sheetName": "Scan App",
        "sourceFile": "xls/Scan.xlsx",
        "featureTreeColumns": ["version", "source", "level2", "level3", "level4", "level5", "level6", "level7"],
        "featureTreeLabels": feature_tree_labels,
        "models": models,
        "quickSetColumns": list(quickset_headers.values()),
        "componentColumns": component_columns,
        "headerStyle": header_style,
        "rows": rows,
    }

    with open(OUT, "w") as f:
        json.dump(out, f, indent=2)
    print(f"Wrote {OUT}: {len(rows)} rows, {len(models)} models")

    with open(SCHEMA) as f:
        schema = json.load(f)
    schema["tabs"] = [t for t in schema["tabs"] if t["name"] != "Scan"]
    schema["tabs"].append({
        "name": "Scan",
        "columns": out["featureTreeColumns"] + [f"models.{m['key']}" for m in models],
    })
    with open(SCHEMA, "w") as f:
        json.dump(schema, f, indent=2)
    print(f"Updated {SCHEMA}")


if __name__ == "__main__":
    main()
