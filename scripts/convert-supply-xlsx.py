#!/usr/bin/env python3
"""Converts xls/Supply.xlsx ('Supply' sheet) into json/data/Supply.json.

Sheet layout (discovered by inspection, not assumed):
  Cols A-E  (1-5):   Version, source, Level 3..5 -- the feature outline.
                     Shallower than Scan/Settings; only 3 tree levels here.
  Cols F-BJ (6-62):  Printer-model columns, no gap before them (unlike
                     Settings' J/K spacer). Same 4-stacked-row header pattern:
                     row1 = family (merged), row3 = segment (merged),
                     row4 = model name, row5 = engine class, row6 = status.
  Col BK (63):        Design Notes, per row. Label on row4, no fill -- same
                     convention as Scan/Settings.
  No Components/Quick Sets/Short-description/Behavior sections on this sheet
  -- a row1 band labeled "EPICS/STORIE..." exists at col BL but has no row4
  sub-label and no data beneath it in any inspected row, so it's treated the
  same way Scan.xlsx's own trailing dead columns were: a formatting leftover,
  not a real column.
Real data ends at row 192; the sheet's max_row (483) and max_col (93) both
include unused trailing formatting artifacts well past the real content.
"""
import json
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

SRC = "../xls/Supply.xlsx"
OUT = "json/data/Supply.json"
SCHEMA = "json/data/_schema.json"

MODEL_START, MODEL_END = column_index_from_string("F"), column_index_from_string("BJ")
NOTES_COL = column_index_from_string("BK")
LAST_ROW = 192

TREE_COLS = [1, 2, 3, 4, 5]
TREE_FIELD_NAMES = ["version", "source", "level3", "level4", "level5"]


def merged_value(ws, row, col):
    cell = ws.cell(row=row, column=col)
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            return ws.cell(row=rng.min_row, column=rng.min_col).value
    return cell.value


def hex_of(cell):
    rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
    if rgb in (None, "00000000"):
        return None
    return f"#{rgb[2:]}" if len(rgb) == 8 else f"#{rgb}"


def merged_fill(ws, row, col):
    cell = ws.cell(row=row, column=col)
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            return hex_of(ws.cell(row=rng.min_row, column=rng.min_col))
    return hex_of(cell)


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def cell_style(cell):
    fill = hex_of(cell)
    bold = bool(cell.font.bold) if cell.font else False
    if fill is None and not bold:
        return None
    return {"fill": fill, "bold": bold}


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Supply"]

    models = []
    seen_names = set()
    for c in range(MODEL_START, MODEL_END + 1):
        name = clean(ws.cell(row=4, column=c).value)
        if not name:
            continue
        # Same "Cherry" reused for two different columns here as in
        # Settings.xlsx -- disambiguate rather than silently lose one
        # column's data to a dict-key collision. See convert-settings-xlsx.py.
        key = name
        if key in seen_names:
            key = f"{name} ({get_column_letter(c)})"
        seen_names.add(key)
        models.append({
            "key": key,
            "column": get_column_letter(c),
            "family": clean(merged_value(ws, 1, c)),
            "familyFill": merged_fill(ws, 1, c),
            "segment": clean(merged_value(ws, 3, c)),
            "segmentFill": merged_fill(ws, 3, c),
            "engineClass": clean(ws.cell(row=5, column=c).value),
            "status": clean(ws.cell(row=6, column=c).value),
        })

    rows = []
    for r in range(7, LAST_ROW + 1):
        level_vals = [clean(ws.cell(row=r, column=c).value) for c in TREE_COLS]
        model_vals = {m["key"]: clean(ws.cell(row=r, column=column_index_from_string(m["column"])).value) for m in models}
        notes = clean(ws.cell(row=r, column=NOTES_COL).value)

        row_has_content = any(level_vals) or any(model_vals.values()) or notes
        if not row_has_content:
            continue

        entry = {"row": r}
        entry.update(dict(zip(TREE_FIELD_NAMES, level_vals)))
        if any(model_vals.values()):
            entry["models"] = {k: v for k, v in model_vals.items() if v is not None}
        if notes:
            entry["designNotes"] = notes

        cell_styles = {}
        for field_name, col in zip(TREE_FIELD_NAMES, TREE_COLS):
            style = cell_style(ws.cell(row=r, column=col))
            if style:
                cell_styles[field_name] = style
        if cell_styles:
            entry["cellStyle"] = cell_styles

        rows.append(entry)

    header_style = {
        "treeHeaderFill": hex_of(ws.cell(row=4, column=1)),
        "statusRowFill": hex_of(ws.cell(row=6, column=MODEL_START)),
        "componentsBandFill": None,
        "componentsBandLabel": None,
        "notesLabel": clean(ws.cell(row=4, column=NOTES_COL).value),
    }

    feature_tree_labels = [clean(ws.cell(row=4, column=c).value) or n for n, c in zip(TREE_FIELD_NAMES, TREE_COLS)]

    out = {
        "tab": "Supply",
        "sheetName": "Supply",
        "sourceFile": "xls/Supply.xlsx",
        "featureTreeColumns": TREE_FIELD_NAMES,
        "featureTreeLabels": feature_tree_labels,
        "models": models,
        "quickSetColumns": [],
        "componentColumns": [],
        "headerStyle": header_style,
        "rows": rows,
    }

    with open(OUT, "w") as f:
        json.dump(out, f, indent=2)
    print(f"Wrote {OUT}: {len(rows)} rows, {len(models)} models")

    with open(SCHEMA) as f:
        schema = json.load(f)
    schema["tabs"] = [t for t in schema["tabs"] if t["name"] != "Supply"]
    schema["tabs"].append({
        "name": "Supply",
        "columns": out["featureTreeColumns"] + [f"models.{m['key']}" for m in models],
    })
    with open(SCHEMA, "w") as f:
        json.dump(schema, f, indent=2)
    print(f"Updated {SCHEMA}")


if __name__ == "__main__":
    main()
