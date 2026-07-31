#!/usr/bin/env python3
"""Converts xls/Settings.xlsx ('Settings' sheet) into json/data/Settings.json.

Sheet layout (discovered by inspection, not assumed):
  Cols A-I  (1-9):   Version, source, Level 3..9 -- the feature outline. Deeper
                     than Scan/2-Line IA (which stop at Level 7), hence
                     FeatureRow gaining level8/level9 fields for this sheet.
                     Row1-2 also carries a "Settings Menu" band label over
                     cols B-I (not captured -- purely a section title with no
                     UI-facing use here, and column A/Version sits outside it).
  Cols J-K  (10-11): Blank spacer columns in the source -- skipped entirely,
                     not carried into the JSON.
  Cols L-BT (12-72): Printer-model columns. Same 4-stacked-row header pattern
                     as Scan: row1 = family (merged), row3 = segment (merged),
                     row4 = model name, row5 = engine class, row6 = status.
                     Family/segment colors vary per model group.
  Cols BU-BX (73-76): "Component: Setting row" -- band on row1, sub-headers
                     "Level 4".."Level 7" on row4 (same pattern as Scan).
  Col BY (77):        "Short description", per row. Band label on row1.
                     Reuses the epicStory/epicLabel slot (that slot is just
                     "one generically-labeled trailing column", not
                     semantically tied to epics -- see lib/iaDocRepo.ts).
  Col BZ (78):        "Behavior", per row. Band label on row1. New
                     behaviorNote/behaviorLabel slot (Settings is the first
                     sheet with a *third* trailing single column).
  Col CA (79):        Design Notes, per row. Label on row4, no fill --
                     matches Scan's notesLabel convention exactly.
Real data ends at row 1393; the sheet's max_row (2393) and max_col (154) both
include unused trailing formatting artifacts well past the real content.
"""
import json
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

SRC = "../xls/Settings.xlsx"
OUT = "json/data/Settings.json"
SCHEMA = "json/data/_schema.json"

MODEL_START, MODEL_END = column_index_from_string("L"), column_index_from_string("BT")
COMPONENT_START, COMPONENT_END = column_index_from_string("BU"), column_index_from_string("BX")
SHORT_DESC_COL = column_index_from_string("BY")
BEHAVIOR_COL = column_index_from_string("BZ")
NOTES_COL = column_index_from_string("CA")
LAST_ROW = 1393

# Real sheet columns are A-I (1-9); J/K (10-11) are a blank spacer and are
# simply not in this list, so they never appear in the JSON.
TREE_COLS = [1, 2, 3, 4, 5, 6, 7, 8, 9]
TREE_FIELD_NAMES = ["version", "source", "level3", "level4", "level5", "level6", "level7", "level8", "level9"]


def merged_value(ws, row, col):
    cell = ws.cell(row=row, column=col)
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            return ws.cell(row=rng.min_row, column=rng.min_col).value
    return cell.value


def hex_of(cell):
    rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
    if rgb in (None, "00000000"):
        return None
    return f"#{rgb[2:]}" if len(rgb) == 8 else f"#{rgb}"


def merged_fill(ws, row, col):
    cell = ws.cell(row=row, column=col)
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            return hex_of(ws.cell(row=rng.min_row, column=rng.min_col))
    return hex_of(cell)


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def cell_style(cell):
    fill = hex_of(cell)
    bold = bool(cell.font.bold) if cell.font else False
    if fill is None and not bold:
        return None
    return {"fill": fill, "bold": bold}


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Settings"]

    models = []
    seen_names = set()
    for c in range(MODEL_START, MODEL_END + 1):
        name = clean(ws.cell(row=4, column=c).value)
        if not name:
            continue
        # The source sheet reuses the model name "Cherry" for two genuinely
        # different columns (N and AR) -- if both used it as their key
        # verbatim, the second would silently overwrite the first everywhere
        # keys are used as dict/lookup identifiers (row.models, colId,
        # React's key prop). Disambiguate rather than lose a column's data;
        # the column letter suffix only appears for entries that actually
        # collide, so unaffected sheets look untouched.
        key = name
        if key in seen_names:
            key = f"{name} ({get_column_letter(c)})"
        seen_names.add(key)
        models.append({
            "key": key,
            "column": get_column_letter(c),
            "family": clean(merged_value(ws, 1, c)),
            "familyFill": merged_fill(ws, 1, c),
            "segment": clean(merged_value(ws, 3, c)),
            "segmentFill": merged_fill(ws, 3, c),
            "engineClass": clean(ws.cell(row=5, column=c).value),
            "status": clean(ws.cell(row=6, column=c).value),
        })

    component_headers = {c: clean(ws.cell(row=4, column=c).value) for c in range(COMPONENT_START, COMPONENT_END + 1)}

    rows = []
    for r in range(7, LAST_ROW + 1):
        level_vals = [clean(ws.cell(row=r, column=c).value) for c in TREE_COLS]
        model_vals = {m["key"]: clean(ws.cell(row=r, column=column_index_from_string(m["column"])).value) for m in models}
        component_vals = {v: clean(ws.cell(row=r, column=c).value) for c, v in component_headers.items() if v}
        short_desc = clean(ws.cell(row=r, column=SHORT_DESC_COL).value)
        behavior = clean(ws.cell(row=r, column=BEHAVIOR_COL).value)
        notes = clean(ws.cell(row=r, column=NOTES_COL).value)

        row_has_content = (
            any(level_vals) or any(model_vals.values()) or any(component_vals.values())
            or short_desc or behavior or notes
        )
        if not row_has_content:
            continue

        entry = {"row": r}
        entry.update(dict(zip(TREE_FIELD_NAMES, level_vals)))
        if any(model_vals.values()):
            entry["models"] = {k: v for k, v in model_vals.items() if v is not None}
        if any(component_vals.values()):
            entry["componentSetting"] = component_vals
        if short_desc:
            entry["epicStory"] = short_desc
        if behavior:
            entry["behaviorNote"] = behavior
        if notes:
            entry["designNotes"] = notes

        cell_styles = {}
        for field_name, col in zip(TREE_FIELD_NAMES, TREE_COLS):
            style = cell_style(ws.cell(row=r, column=col))
            if style:
                cell_styles[field_name] = style
        if cell_styles:
            entry["cellStyle"] = cell_styles

        rows.append(entry)

    header_style = {
        "treeHeaderFill": hex_of(ws.cell(row=4, column=1)),
        "statusRowFill": hex_of(ws.cell(row=6, column=MODEL_START)),
        "componentsBandFill": hex_of(ws.cell(row=1, column=COMPONENT_START)),
        "componentsBandLabel": clean(ws.cell(row=1, column=COMPONENT_START).value),
        "epicBandFill": hex_of(ws.cell(row=1, column=SHORT_DESC_COL)),
        "epicLabel": clean(ws.cell(row=1, column=SHORT_DESC_COL).value),
        "behaviorBandFill": hex_of(ws.cell(row=1, column=BEHAVIOR_COL)),
        "behaviorLabel": clean(ws.cell(row=1, column=BEHAVIOR_COL).value),
        "notesLabel": clean(ws.cell(row=4, column=NOTES_COL).value),
    }

    feature_tree_labels = [clean(ws.cell(row=4, column=c).value) or n for n, c in zip(TREE_FIELD_NAMES, TREE_COLS)]
    component_columns = [v for v in component_headers.values() if v]

    out = {
        "tab": "Settings",
        "sheetName": "Settings",
        "sourceFile": "xls/Settings.xlsx",
        "featureTreeColumns": TREE_FIELD_NAMES,
        "featureTreeLabels": feature_tree_labels,
        "models": models,
        "quickSetColumns": [],
        "componentColumns": component_columns,
        "headerStyle": header_style,
        "rows": rows,
    }

    with open(OUT, "w") as f:
        json.dump(out, f, indent=2)
    print(f"Wrote {OUT}: {len(rows)} rows, {len(models)} models")

    with open(SCHEMA) as f:
        schema = json.load(f)
    schema["tabs"] = [t for t in schema["tabs"] if t["name"] != "Settings"]
    schema["tabs"].append({
        "name": "Settings",
        "columns": out["featureTreeColumns"] + [f"models.{m['key']}" for m in models],
    })
    with open(SCHEMA, "w") as f:
        json.dump(schema, f, indent=2)
    print(f"Updated {SCHEMA}")


if __name__ == "__main__":
    main()
