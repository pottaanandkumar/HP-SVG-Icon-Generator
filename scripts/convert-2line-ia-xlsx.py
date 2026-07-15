#!/usr/bin/env python3
"""Converts xls/2-Line IA.xlsx into json/data/2-Line IA.json.

Same shape as convert-scan-xlsx.py's source sheet (feature outline + model
matrix + a "Components: Setting row" side-table), but smaller and with
different header row positions -- discovered by inspection, not assumed:
  Cols A-H (1-8):  Version, source, Level 1..6 -- the feature outline.
                   Mapped positionally into FeatureRow's level2..level7
                   fields (that type's field names come from the Scan sheet,
                   but the shape -- 6 level columns -- matches here too).
  Cols I-J (9-10): 2 model columns. Header is stacked: row2 = family
                   (merged I2:J3, "2-Line"), row4 = segment (merged I4:J4,
                   "POLESTAR"), row5 = model name, row6 = WIP/Ready status.
                   No engine-class row in this sheet (unlike Scan).
  Cols L-Q (12-17): "Components: Setting row" -- per-row UI widget metadata,
                   headers "Level 1".."Level 6" on row 4.
No Factory Quick Sets / EPICS columns in this sheet.
"""
import json
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

SRC = "../xls/2-Line IA.xlsx"
TAB_NAME = "2-Line IA"
OUT = f"json/data/{TAB_NAME}.json"
SCHEMA = "json/data/_schema.json"

MODEL_START, MODEL_END = column_index_from_string("I"), column_index_from_string("J")
COMPONENT_START, COMPONENT_END = column_index_from_string("L"), column_index_from_string("Q")


def merged_value(ws, row, col):
    cell = ws.cell(row=row, column=col)
    for rng in ws.merged_cells.ranges:
        if cell.coordinate in rng:
            return ws.cell(row=rng.min_row, column=rng.min_col).value
    return cell.value


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def cell_style(cell):
    """Captures the original xlsx presentation for one cell -- fill color and
    bold -- so the frontend can reproduce the sheet's own visual tree
    language (see module docstring) instead of re-deriving depth from
    scratch. Returns None for a plain/unstyled cell to keep the JSON small."""
    fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
    if fill in (None, "00000000"):
        fill = None
    else:
        fill = f"#{fill[2:]}" if len(fill) == 8 else f"#{fill}"
    bold = bool(cell.font.bold) if cell.font else False
    if fill is None and not bold:
        return None
    return {"fill": fill, "bold": bold}


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[TAB_NAME]

    last_row = 0
    for r in range(1, ws.max_row + 1):
        if any(clean(ws.cell(row=r, column=c).value) is not None for c in range(1, COMPONENT_END + 1)):
            last_row = r

    models = []
    for c in range(MODEL_START, MODEL_END + 1):
        name = clean(ws.cell(row=5, column=c).value)
        if not name:
            continue
        models.append({
            "key": name,
            "column": get_column_letter(c),
            "family": clean(merged_value(ws, 2, c)),
            "segment": clean(merged_value(ws, 4, c)),
            "engineClass": None,
            "status": clean(ws.cell(row=6, column=c).value),
        })

    component_headers = {c: clean(ws.cell(row=4, column=c).value) for c in range(COMPONENT_START, COMPONENT_END + 1)}

    level_field_names = ["level2", "level3", "level4", "level5", "level6", "level7"]

    style_field_names = ["version", "source"] + level_field_names

    rows = []
    for r in range(7, last_row + 1):
        version = clean(ws.cell(row=r, column=1).value)
        source = clean(ws.cell(row=r, column=2).value)
        levels = [clean(ws.cell(row=r, column=c).value) for c in range(3, 9)]
        model_vals = {m["key"]: clean(ws.cell(row=r, column=column_index_from_string(m["column"])).value) for m in models}
        component_vals = {v: clean(ws.cell(row=r, column=c).value) for c, v in component_headers.items() if v}

        row_has_content = version or source or any(levels) or any(model_vals.values()) or any(component_vals.values())
        if not row_has_content:
            continue

        entry = {"row": r, "version": version, "source": source}
        entry.update(dict(zip(level_field_names, levels)))
        if any(model_vals.values()):
            entry["models"] = {k: v for k, v in model_vals.items() if v is not None}
        if any(component_vals.values()):
            entry["componentSetting"] = component_vals

        # Original xlsx presentation (fill color, bold) for the feature-tree
        # columns on this row -- this is what lets the frontend render the
        # sheet's own tree/depth language faithfully instead of flattening
        # it. Cols 1-2 = version/source, cols 3-8 = Level 1-6.
        cell_styles = {}
        for field_name, col in zip(style_field_names, range(1, 9)):
            style = cell_style(ws.cell(row=r, column=col))
            if style:
                cell_styles[field_name] = style
        if cell_styles:
            entry["cellStyle"] = cell_styles

        rows.append(entry)

    def hex_of(cell):
        rgb = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
        if rgb in (None, "00000000"):
            return None
        return f"#{rgb[2:]}" if len(rgb) == 8 else f"#{rgb}"

    header_style = {
        "treeHeaderFill": hex_of(ws.cell(row=5, column=1)),
        "modelHeaderFill": hex_of(ws.cell(row=5, column=MODEL_START)),
        "modelSegmentFill": hex_of(ws.cell(row=4, column=MODEL_START)),
        "statusRowFill": hex_of(ws.cell(row=6, column=1)),
        "componentsBandFill": hex_of(ws.cell(row=1, column=COMPONENT_START)),
        "componentsBandLabel": clean(ws.cell(row=1, column=COMPONENT_START).value),
    }

    feature_tree_labels = [clean(ws.cell(row=5, column=c).value) or n for n, c in zip(style_field_names, range(1, 9))]

    # Ordered component column labels ("Level 1".."Level 6"), matching the
    # keys used inside each row's componentSetting object -- lets the
    # frontend render these as real columns instead of the flattened,
    # potentially-missing lookup it was doing before.
    component_columns = [v for v in component_headers.values() if v]

    out = {
        "tab": TAB_NAME,
        "sheetName": TAB_NAME,
        "sourceFile": "xls/2-Line IA.xlsx",
        "featureTreeColumns": ["version", "source"] + level_field_names,
        "featureTreeLabels": feature_tree_labels,
        "models": models,
        "quickSetColumns": [],
        "componentColumns": component_columns,
        "headerStyle": header_style,
        "rows": rows,
    }

    with open(OUT, "w") as f:
        json.dump(out, f, indent=2)
    print(f"Wrote {OUT}: {len(rows)} rows, {len(models)} models")

    with open(SCHEMA) as f:
        schema = json.load(f)
    schema["tabs"] = [t for t in schema["tabs"] if t["name"] != TAB_NAME]
    schema["tabs"].append({
        "name": TAB_NAME,
        "columns": out["featureTreeColumns"] + [f"models.{m['key']}" for m in models],
    })
    with open(SCHEMA, "w") as f:
        json.dump(schema, f, indent=2)
    print(f"Updated {SCHEMA}")


if __name__ == "__main__":
    main()
